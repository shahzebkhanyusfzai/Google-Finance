import openpyxl
import undetected_chromedriver as uc
from selenium_stealth import stealth
from webdriver_manager.chrome import ChromeDriverManager
from scrapy import Selector
from selenium.webdriver.chrome.service import Service
import time
from selenium.webdriver.support.ui import Select, WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import re
from selenium.webdriver.chrome.options import Options

 
def convert_to_number(string):
    if 'M' in string:
        return float(string.replace('M', '')) * 1e6
    elif 'K' in string:
        return float(string.replace('K', '')) * 1e3
    elif 'B' in string:
        return float(string.replace('B', '')) * 1e9
    elif 'T' in string:  # Handle trillion
        return float(string.replace('T', '')) * 1e12
    else:
        return float(string)
    
def extract_value(s):
    match = None
    if s is not None:
        match = re.match(r"(-?[\d\.]+[KMBT])", s) 
    if match:
        return match.group(1)
    return None

def getdriver():
    options = Options()
    # options.headless = True
    options.page_load_strategy = 'normal'
    driver = webdriver.Chrome(options=options, service=Service(ChromeDriverManager().install()))
    return driver

def get_tickers_from_excel(file_path):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb["Updated Raw"]
    tickers = [cell.value for row in sheet.iter_rows(min_row=355, min_col=3, max_col=3) for cell in row if cell.value is not None]
    return tickers

def update_excel_with_data(file_path, data, ticker):
    print('data writing')
    wb = openpyxl.load_workbook(file_path)
    sheet = wb["Updated Raw"]

    mapping = {
        "market": "D", # Market
        "founded": "E", # Year Founded
        "currency": "M", # Currency
        "Price": "N", # Price
        "Earnings Per Share": "O",
        "P/E Ratio": "P", #
        "Dividend_Yield": "Q",
        "Earnings_per_share_2021": "R",
        'price_2022':'W',    
        "Earnings Per Share": "Y",
        'price_2021': 'AC',
        "Earnings_per_share_2020": "AD",
        'price_2020': 'AG',
        'Earnings_per_share_2019': 'AH',
        'price_2019':'AK',
        'Earnings_per_share_2018': 'AL',
        "Revenue": "AO", 
        "Ebitda": "AQ", 
        "Net_income": "AS", 
        "Net_Profit_margin": "AU", 
        "Return_On_Assets": "AV", 
        "Price_to_book": "AX", 
        "market_Cap": "AY", 
        "Outstanding_Shares": "BA", 
        "Average_daily_vol": "BB", 
    }

    convert_keys = ['market_Cap', 'Ebitda', 'Revenue', 'Net_income','Outstanding_Shares','Average_daily_vol']

    for row_num, cell in enumerate(sheet['C'], 1):  # Starting from column 'C' as tickers are there

        if cell.value is not None and cell.value == ticker and row_num >= 355:
            for key, column in mapping.items():
                value_to_write = data.get(key, None)

                if key in convert_keys and value_to_write is not None:
                    try:
                        value_to_write = convert_to_number(value_to_write)
                    except ValueError:
                        print(f"Failed to convert {value_to_write} for key {key}")

                sheet[f"{column}{row_num}"].value = value_to_write # Using .get() to avoid KeyError
                print(f"ticker: {ticker} is written to excel")
            wb.save(file_path)  # Don't forget to save your changes
            break  # Exit the loop once match found
        
def scraping_time_series_graph(driver):
    data_points = []
    time.sleep(4)
    try:
        graph = driver.find_element(By.XPATH,"//*[name()='svg']/*[name()='g']/descendant::*[name()='g'][@class='gJBfM']")
    except:
        pass
    time.sleep(2)
    try:
        for x in range(-325, 325):
            action = ActionChains(driver).move_to_element_with_offset(graph, x, graph.size['height'] / 2)
            action.perform()
            response = Selector(text=driver.page_source)
            price = response.xpath("//div[@class='hSGhwc']/p[@jsname='BYCTfd']/text()").get()
            date = response.xpath("//div[@class='hSGhwc']/p[@jsname='LlMULe']/text()").get()
            volume = response.xpath("//div[@class='hSGhwc']/p[@jsname='R30goc']/span/text()").get()
            
            data = {
                'price': price,
                'date': date,
                'volume': volume
            }
            data_points.append(data)
        return data_points
    except:
        data_points = ''

def getting_yearly_price_from_Scraping_time_series_func(data_points):
    for data in data_points:
        if data['price'] is not None:
            price_at_list = data['price']
            break


    target_years = ['2022','2021', '2020', '2019']
    prices_for_target_years = {}
    months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

    # Initialize the dictionary with the start of each target year for comparison purposes
    for year in target_years:
        prices_for_target_years[year] = {
            'date': (int(year), 1, 1),
            'price': None
        }

    for data in data_points:
        if data['date']:
            data['date'] = data['date'].replace(',', ' ')
            month_str, day, year = data['date'].split()

            month_num = months.index(month_str) + 1

            # Create a tuple representation of the date for easy comparison
            date_tuple = (int(year), month_num, int(day))

        # If this year is one of our target years
            if year in target_years:
                # If we haven't stored a price for this year yet, or this date is later than the stored one
                if date_tuple > prices_for_target_years[year]['date']:
                    # Store the data in our results dictionary
                    prices_for_target_years[year] = {
                        'date': date_tuple,
                        'price': data['price']
                }

# Extracting prices for the latest dates in each year
    final_prices = {year: data['price'] for year, data in prices_for_target_years.items()}
    final_prices['list_price'] = price_at_list
    print(final_prices)
    return final_prices

def rearrange_string(text):
  parts = text.split(':')
  symbol = parts[1].strip()
  exchange = parts[0].strip()
  return f"{symbol}:{exchange}"


